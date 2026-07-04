from __future__ import annotations

import csv
import hashlib
import re
from datetime import date, datetime
from io import BytesIO, TextIOWrapper
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ALLOWED_EXTENSIONS = {".xlsx", ".csv"}
MAX_FILE_SIZE = 20 * 1024 * 1024


def stable_field_id(name: str, index: int) -> str:
    normalized = re.sub(r"\W+", "_", name.strip().lower()).strip("_") or f"column_{index + 1}"
    return f"field_{normalized}_{index + 1}"


def infer_type(values: list[Any]) -> str:
    samples = [value for value in values if value not in (None, "")][:100]
    if not samples:
        return "text"
    if all(isinstance(value, bool) for value in samples):
        return "boolean"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in samples):
        return "integer"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in samples):
        return "number"
    if all(isinstance(value, (datetime, date)) for value in samples):
        return "datetime"
    text = [str(value).strip() for value in samples]
    if all(re.fullmatch(r"-?\d+(\.\d+)?%", value) for value in text):
        return "percentage"
    if all(re.fullmatch(r"-?\d+(\.\d+)?", value.replace(",", "")) for value in text):
        return "number"
    return "text"


def build_sheet(name: str, rows: list[tuple[Any, ...]] | list[list[str]], preview_rows: int) -> dict[str, Any]:
    if not rows:
        return {"name": name, "columns": [], "preview": [], "row_count": 0}
    headers = [str(value) if value not in (None, "") else f"Column {index + 1}" for index, value in enumerate(rows[0])]
    data_rows = rows[1:]
    columns = []
    for index, header in enumerate(headers):
        values = [row[index] if index < len(row) else None for row in data_rows]
        columns.append({"id": stable_field_id(header, index), "name": header, "type": infer_type(values), "index": index})
    preview = [dict(zip(headers, row, strict=False)) for row in data_rows[:preview_rows]]
    return {"name": name, "columns": columns, "preview": preview, "row_count": len(data_rows)}


def inspect_file(filename: str, content: bytes, preview_rows: int = 20) -> dict[str, Any]:
    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValueError("当前版本仅支持 .xlsx 和 .csv 文件")
    if not content:
        raise ValueError("上传文件为空")
    if len(content) > MAX_FILE_SIZE:
        raise ValueError("上传文件超过 20 MB 限制")
    if extension == ".csv":
        stream = TextIOWrapper(BytesIO(content), encoding="utf-8-sig", newline="")
        sheets = [build_sheet("CSV", list(csv.reader(stream)), preview_rows)]
    else:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheets = [build_sheet(sheet.title, list(sheet.iter_rows(values_only=True)), preview_rows) for sheet in workbook.worksheets]
    return {"filename": filename, "size": len(content), "checksum": hashlib.sha256(content).hexdigest(), "sheets": sheets}


def compare_schemas(previous: list[dict[str, Any]], current: list[dict[str, Any]]) -> dict[str, Any]:
    old = {column["id"]: column for sheet in previous for column in sheet.get("columns", [])}
    new = {column["id"]: column for sheet in current for column in sheet.get("columns", [])}
    added = [new[key] for key in new.keys() - old.keys()]
    removed = [old[key] for key in old.keys() - new.keys()]
    changed = [{"previous": old[key], "current": new[key]} for key in old.keys() & new.keys() if old[key]["type"] != new[key]["type"]]
    return {"added": added, "removed": removed, "changed": changed, "compatible": not removed and not changed}
